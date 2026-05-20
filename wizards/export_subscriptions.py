import base64
import io
from odoo import models, fields, api, _
from odoo.exceptions import UserError


class NahjExportSubscriptionsWizard(models.TransientModel):
    """Export subscription data to an Excel workbook."""
    _name = 'nahj.export.subscriptions.wizard'
    _description = 'Export Subscriptions to Excel'

    tower_id = fields.Many2one(
        'nahj.tower', string='Tower',
        help='Leave empty to export all towers.')
    year = fields.Integer(
        string='Year',
        default=lambda self: fields.Date.today().year,
        help='Leave 0 to export all years.')
    month = fields.Selection([
        ('0',  'All Months'),
        ('1',  'January'),  ('2',  'February'), ('3',  'March'),
        ('4',  'April'),    ('5',  'May'),       ('6',  'June'),
        ('7',  'July'),     ('8',  'August'),    ('9',  'September'),
        ('10', 'October'),  ('11', 'November'),  ('12', 'December'),
    ], string='Month', default='0')
    state = fields.Selection([
        ('all',    'All'),
        ('paid',   'Paid Only'),
        ('unpaid', 'Unpaid Only'),
    ], string='Status', default='all')

    # ── Result fields ─────────────────────────────────────────────────────────
    export_file = fields.Binary(string='Download File', readonly=True)
    export_filename = fields.Char(string='File Name', readonly=True)
    result_state = fields.Selection([
        ('draft', 'Ready'),
        ('done',  'Done'),
    ], default='draft')

    def action_export(self):
        self.ensure_one()

        # ── Build domain ──────────────────────────────────────────────────────
        domain = []
        if self.tower_id:
            domain.append(('tower_id', '=', self.tower_id.id))
        if self.year:
            domain.append(('year', '=', self.year))
        if self.month and self.month != '0':
            domain.append(('month', '=', self.month))
        if self.state != 'all':
            domain.append(('state', '=', self.state))

        records = self.env['nahj.subscription'].search(
            domain, order='year desc, month desc, tower_id, unit_id')
        if not records:
            raise UserError(_('No records found for the selected filters.'))

        # ── Build Excel workbook ──────────────────────────────────────────────
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            HAS_OPENPYXL = True
        except ImportError:
            HAS_OPENPYXL = False

        if HAS_OPENPYXL:
            output = self._build_xlsx(records, openpyxl)
            filename = 'subscriptions_export.xlsx'
            mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            output = self._build_csv(records)
            filename = 'subscriptions_export.csv'

        self.write({
            'export_file': base64.b64encode(output),
            'export_filename': filename,
            'result_state': 'done',
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'views': [[False, 'form']],
            'target': 'new',
        }

    # ── Excel builder ─────────────────────────────────────────────────────────
    def _build_xlsx(self, records, openpyxl):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Subscriptions'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='1A3C5E')
        center = Alignment(horizontal='center', vertical='center')
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'))

        headers = [
            'Tower', 'Unit', 'Floor', 'Resident', 'Phone',
            'Year', 'Month', 'Amount', 'Status',
            'Late Fee', 'Payment Date', 'Payment Method', 'Notes',
        ]
        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin

        month_names = {
            '1': 'January', '2': 'February', '3': 'March',
            '4': 'April', '5': 'May', '6': 'June',
            '7': 'July', '8': 'August', '9': 'September',
            '10': 'October', '11': 'November', '12': 'December',
        }
        paid_fill   = PatternFill('solid', fgColor='E8F5E9')
        unpaid_fill = PatternFill('solid', fgColor='FFEBEE')

        for row_idx, sub in enumerate(records, 2):
            resident = sub.unit_id.resident_id
            row_data = [
                sub.tower_id.name or '',
                sub.unit_id.name or '',
                sub.unit_id.floor or 0,
                resident.name if resident else '',
                resident.phone if resident else '',
                sub.year,
                month_names.get(sub.month, sub.month),
                sub.amount,
                'Paid' if sub.state == 'paid' else 'Unpaid',
                sub.late_fee_amount or 0.0,
                str(sub.payment_date) if sub.payment_date else '',
                dict(sub._fields['payment_method'].selection).get(
                    sub.payment_method, '') if sub.payment_method else '',
                sub.notes or '',
            ]
            ws.append(row_data)
            row_fill = paid_fill if sub.state == 'paid' else unpaid_fill
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = row_fill
                cell.border = thin
                cell.alignment = Alignment(vertical='center')

        # Column widths
        col_widths = [20, 10, 8, 25, 15, 8, 12, 14, 10, 12, 14, 18, 30]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)].width = w

        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── CSV fallback ──────────────────────────────────────────────────────────
    def _build_csv(self, records):
        import csv
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow([
            'Tower', 'Unit', 'Floor', 'Resident', 'Phone',
            'Year', 'Month', 'Amount', 'Status',
            'Late Fee', 'Payment Date', 'Payment Method', 'Notes',
        ])
        for sub in records:
            resident = sub.unit_id.resident_id
            writer.writerow([
                sub.tower_id.name or '',
                sub.unit_id.name or '',
                sub.unit_id.floor or 0,
                resident.name if resident else '',
                resident.phone if resident else '',
                sub.year,
                sub.month,
                sub.amount,
                sub.state,
                sub.late_fee_amount or 0.0,
                str(sub.payment_date) if sub.payment_date else '',
                sub.payment_method or '',
                sub.notes or '',
            ])
        return buf.getvalue().encode('utf-8-sig')
