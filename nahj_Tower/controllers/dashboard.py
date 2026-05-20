import io
import json
import calendar as _calendar
from datetime import date, timedelta
from odoo import http, _
from odoo.exceptions import AccessError
from odoo.http import request, Response


def _json(data):
    return Response(
        json.dumps(data, default=str),
        mimetype='application/json',
    )


class NahjDashboardController(http.Controller):

    @http.route('/nahj_tower/dashboard_data', type='json', auth='user')
    def dashboard_data(self, tower_id=None, year=None, month=None,
                       date_from=None, date_to=None):
        # Menu is restricted to Nahj groups, but JSON routes are callable by any
        # logged-in user — enforce the same groups and never use sudo() here.
        if not request.env.user.has_group('nahj_Tower.group_nahj_user'):
            raise AccessError(_(
                'You do not have access to the Nahj Tower dashboard.\n'
                'ليس لديك صلاحية الوصول إلى لوحة تحكم نهج تاور.'))
        env = request.env
        today = date.today()
        try:
            selected_year  = int(year)  if year  else None
            selected_month = int(month) if month else None
        except (ValueError, TypeError):
            selected_year = selected_month = None

        # Parse custom date range
        from_date = to_date = None
        if date_from:
            try: from_date = date.fromisoformat(str(date_from))
            except Exception: pass
        if date_to:
            try: to_date = date.fromisoformat(str(date_to))
            except Exception: pass

        # ── Domain ────────────────────────────────────────────────────────────
        tower_domain = []
        if tower_id not in (None, '', False):
            try:
                tid = int(tower_id)
                if tid:
                    tower_domain = [('id', '=', tid)]
            except (TypeError, ValueError):
                tower_domain = [('id', '=', -1)]
        towers = env['nahj.tower'].search(tower_domain)

        # ── KPI Cards ─────────────────────────────────────────────────────────
        all_units = env['nahj.unit'].search(
            [('tower_id', 'in', towers.ids)] if towers else [])

        base_sub = [('tower_id', 'in', towers.ids)] if towers else []
        base_exp = ([('tower_id', 'in', towers.ids)] if towers else []) + [
            ('state', '=', 'confirmed')]

        # ── Two filtering modes ────────────────────────────────────────────────
        # Mode A – Date-range (today/week/quarter…): paid subs filtered by
        #   payment_date (actual cash received); expenses filtered by their date.
        # Mode B – Period (year ± month): filter by subscription billing period.
        is_date_range = bool(from_date and to_date)

        if is_date_range:
            exp_domain = base_exp + [
                ('date', '>=', from_date.isoformat()),
                ('date', '<=', to_date.isoformat()),
            ]
            # Paid: only subs with payment_date inside the window
            paid_subs = env['nahj.subscription'].search(
                base_sub + [
                    ('state', '=', 'paid'),
                    ('payment_date', '>=', from_date.isoformat()),
                    ('payment_date', '<=', to_date.isoformat()),
                ]
            )
            # Unpaid: all outstanding (needed for overdue / collection-rate)
            unpaid_subs = env['nahj.subscription'].search(
                base_sub + [('state', '=', 'unpaid')])
            all_subs = paid_subs | unpaid_subs
        else:
            sub_domain = list(base_sub)
            exp_domain = list(base_exp)
            if selected_year:
                sub_domain.append(('year', '=', str(selected_year)))
                exp_domain += [
                    ('date', '>=', '%d-01-01' % selected_year),
                    ('date', '<=', '%d-12-31' % selected_year),
                ]
                if selected_month:
                    last_day = _calendar.monthrange(selected_year, selected_month)[1]
                    sub_domain.append(('month', '=', str(selected_month)))
                    exp_domain = [d for d in exp_domain if not (
                        isinstance(d, (list, tuple)) and len(d) == 3 and d[0] == 'date')]
                    exp_domain += [
                        ('date', '>=', '%d-%02d-01' % (selected_year, selected_month)),
                        ('date', '<=', '%d-%02d-%02d' % (
                            selected_year, selected_month, last_day)),
                    ]
            all_subs = env['nahj.subscription'].search(sub_domain)
            paid_subs = all_subs.filtered(lambda s: s.state == 'paid')
            unpaid_subs = all_subs.filtered(lambda s: s.state == 'unpaid')

        all_expenses = env['nahj.expense'].search(exp_domain)

        # Use amount_after_discount – the actual charged amount after unit-level discounts
        total_subscriptions = sum(paid_subs.mapped('amount_after_discount'))
        total_expenses = sum(all_expenses.mapped('amount'))
        balance = total_subscriptions - total_expenses

        # ── Overdue (always global – independent of the active filter) ─────────
        # Current month is NOT overdue; only strictly past unpaid months are.
        all_unpaid_global = env['nahj.subscription'].search(
            base_sub + [('state', '=', 'unpaid')])
        overdue_units = set()
        for sub in all_unpaid_global:
            try:
                sub_yr = int(sub.year)
                sub_mo = int(sub.month)
            except (TypeError, ValueError):
                continue
            if (sub_yr < today.year or
                    (sub_yr == today.year and sub_mo < today.month)):
                overdue_units.add(sub.unit_id.id)

        # ── "This Period" Summary (Row 2 KPIs) ────────────────────────────────
        if is_date_range or selected_month:
            # Date-range or specific month → period data = already-filtered set
            period_paid     = paid_subs
            period_unpaid   = unpaid_subs
            period_expenses = all_expenses
        elif selected_year:
            # Year view → show current calendar month within the selected year
            _cur_m = today.month
            _p = all_subs.filtered(lambda s: int(s.month) == _cur_m)
            period_paid     = _p.filtered(lambda s: s.state == 'paid')
            period_unpaid   = _p.filtered(lambda s: s.state == 'unpaid')
            period_expenses = all_expenses.filtered(
                lambda e: e.date and e.date.month == _cur_m)
        else:
            # No filter → current month of current year
            _p = all_subs.filtered(
                lambda s: s.year == str(today.year) and s.month == str(today.month))
            period_paid     = _p.filtered(lambda s: s.state == 'paid')
            period_unpaid   = _p.filtered(lambda s: s.state == 'unpaid')
            period_expenses = all_expenses.filtered(
                lambda e: e.date and
                          e.date.year == today.year and
                          e.date.month == today.month)

        # ── Daily Cash Flow – last 7 days ─────────────────────────────────────
        daily_flow = []
        for i in range(6, -1, -1):
            day = today - timedelta(days=i)
            day_subs = paid_subs.filtered(
                lambda s, d=day: s.payment_date == d)
            day_expenses = all_expenses.filtered(
                lambda e, d=day: e.date == d)
            daily_flow.append({
                'label': day.strftime('%a'),
                'date': day.isoformat(),
                'income': sum(day_subs.mapped('amount_after_discount')),
                'expense': sum(day_expenses.mapped('amount')),
            })

        # ── Expense Breakdown by Category ─────────────────────────────────────
        expense_by_cat = {}
        for exp in all_expenses:
            cat = exp.category_id.name if exp.category_id else 'Other'
            expense_by_cat[cat] = expense_by_cat.get(cat, 0) + exp.amount
        expense_chart = [
            {'label': k, 'value': v}
            for k, v in sorted(expense_by_cat.items(),
                               key=lambda x: x[1], reverse=True)
        ]

        # ── Subscription Status per Tower ─────────────────────────────────────
        tower_stats = []
        for t in towers:
            t_subs = all_subs.filtered(lambda s: s.tower_id.id == t.id)
            t_paid = t_subs.filtered(lambda s: s.state == 'paid')
            t_unpaid = t_subs.filtered(lambda s: s.state == 'unpaid')
            tower_stats.append({
                'id': t.id,
                'name': t.name,
                'unit_count': len(t.unit_ids),
                'occupied': len(t.unit_ids.filtered(
                    lambda u: u.state == 'occupied')),
                'vacant': len(t.unit_ids.filtered(
                    lambda u: u.state == 'vacant')),
                'paid_amount': sum(t_paid.mapped('amount_after_discount')),
                'unpaid_amount': sum(t_unpaid.mapped('amount_after_discount')),
                'expenses': sum(
                    all_expenses.filtered(
                        lambda e: e.tower_id.id == t.id).mapped('amount')),
                'balance': sum(t_paid.mapped('amount_after_discount')) - sum(
                    all_expenses.filtered(
                        lambda e: e.tower_id.id == t.id).mapped('amount')),
            })

        # ── Recent Overdue Subscriptions (past months only) ───────────────────
        # Current month unpaid is NOT overdue – already shown in the monthly KPI row
        def _is_overdue(s):
            try:
                y, m = int(s.year), int(s.month)
            except (TypeError, ValueError):
                return False
            return y < today.year or (y == today.year and m < today.month)

        def _overdue_sort_key(s):
            try:
                return (int(s.year), int(s.month))
            except (TypeError, ValueError):
                return (0, 0)

        truly_overdue = unpaid_subs.filtered(_is_overdue)
        recent_unpaid = truly_overdue.sorted(key=_overdue_sort_key, reverse=True)[:10]
        recent_unpaid_data = [{
            'unit': '%s – %s' % (s.tower_id.name, s.unit_id.name),
            'resident': s.unit_id.resident_id.name or '—',
            'month': s.month,
            'year': s.year,
            'amount': s.amount_after_discount,
            'id': s.id,
        } for s in recent_unpaid]

        # ── Recent Expenses ────────────────────────────────────────────────────
        recent_expenses = all_expenses.sorted(key=lambda e: e.date or date.min,
                                              reverse=True)[:8]
        recent_expenses_data = [{
            'tower': e.tower_id.name,
            'name': e.name,
            'category': e.category_id.name if e.category_id else '—',
            'date': e.date.isoformat() if e.date else '',
            'amount': e.amount,
        } for e in recent_expenses]

        # ── Payment Method Breakdown ──────────────────────────────────────────
        payment_methods = [
            ('cash',     'نقدي / Cash',          '#27ae60'),
            ('transfer', 'تحويل / Transfer',      '#2980b9'),
            ('check',    'شيك / Check',           '#8e44ad'),
            ('other',    'أخرى / Other',          '#7f8c8d'),
        ]
        payment_method_stats = []
        for code, label, color in payment_methods:
            method_subs = paid_subs.filtered(
                lambda s, c=code: s.payment_method == c)
            if method_subs:
                payment_method_stats.append({
                    'code':   code,
                    'label':  label,
                    'color':  color,
                    'count':  len(method_subs),
                    'amount': sum(method_subs.mapped('amount_after_discount')),
                })
        # subs paid with no method set
        no_method_subs = paid_subs.filtered(lambda s: not s.payment_method)
        if no_method_subs:
            payment_method_stats.append({
                'code':   'unset',
                'label':  'غير محدد / Unset',
                'color':  '#bdc3c7',
                'count':  len(no_method_subs),
                'amount': sum(no_method_subs.mapped('amount_after_discount')),
            })

        return {
            # KPIs
            'total_towers': len(towers),
            'total_units': len(all_units),
            'occupied_units': len(all_units.filtered(
                lambda u: u.state == 'occupied')),
            'vacant_units': len(all_units.filtered(
                lambda u: u.state == 'vacant')),
            'total_subscriptions': total_subscriptions,
            'total_expenses': total_expenses,
            'balance': balance,
            'overdue_count': len(overdue_units),
            'collection_rate': round(
                total_subscriptions / (total_subscriptions + sum(
                    unpaid_subs.mapped('amount_after_discount')) or 1) * 100, 1),
            # This period (active filter, or current month when viewing a full year)
            'month_paid_count': len(period_paid),
            'month_unpaid_count': len(period_unpaid),
            'month_paid_amount': sum(period_paid.mapped('amount_after_discount')),
            'month_unpaid_amount': sum(period_unpaid.mapped('amount_after_discount')),
            'month_expense_amount': sum(period_expenses.mapped('amount')),
            # Charts & Tables
            'daily_flow': daily_flow,
            'expense_chart': expense_chart,
            'tower_stats': tower_stats,
            'recent_unpaid': recent_unpaid_data,
            'recent_expenses': recent_expenses_data,
            # Payment methods
            'payment_method_stats': payment_method_stats,
            # Filters
            'towers_list': [{'id': t.id, 'name': t.name} for t in
                            env['nahj.tower'].search([])],
            'selected_year': selected_year or today.year,
            'available_years': list(range(today.year, today.year - 6, -1)),
        }

    # ── Import Template Download ───────────────────────────────────────────────
    @http.route('/nahj_tower/import_template', type='http', auth='user')
    def import_template(self):
        """Return a styled .xlsx template (falls back to CSV if openpyxl missing)."""
        if not request.env.user.has_group('nahj_Tower.group_nahj_manager'):
            raise AccessError(_(
                'Only Nahj Tower Managers can download the import template.\n'
                'يُسمح لمديري نهج تاور فقط بتنزيل قالب الاستيراد.'))
        headers_en = ['unit_no', 'floor', 'monthly_subscription',
                      'resident_name', 'resident_phone']
        sample_rows = [
            ['101', '1', '500', 'أحمد محمد', '0501234567'],
            ['102', '1', '500', 'سارة علي',  '0507654321'],
            ['201', '2', '600', '',           ''],
        ]

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Units'

            # Header row
            for col, header in enumerate(headers_en, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font      = Font(bold=True, color='FFFFFF', name='Cairo')
                cell.fill      = PatternFill('solid', fgColor='1A2942')
                cell.alignment = Alignment(horizontal='center')

            # Sample data
            for row_idx, row_data in enumerate(sample_rows, start=2):
                for col_idx, val in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=val)

            # Column widths
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 22

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            return request.make_response(
                buf.read(),
                headers=[
                    ('Content-Type',
                     'application/vnd.openxmlformats-officedocument'
                     '.spreadsheetml.sheet'),
                    ('Content-Disposition',
                     'attachment; filename="nahj_tower_import_template.xlsx"'),
                ],
            )

        except ImportError:
            # Fallback to CSV
            import csv as csv_mod
            buf = io.StringIO()
            writer = csv_mod.writer(buf)
            writer.writerow(headers_en)
            writer.writerows(sample_rows)
            return request.make_response(
                buf.getvalue().encode('utf-8'),
                headers=[
                    ('Content-Type', 'text/csv; charset=utf-8'),
                    ('Content-Disposition',
                     'attachment; filename="nahj_tower_import_template.csv"'),
                ],
            )
